import tempfile
from pathlib import Path
from typing import List
from collections import OrderedDict

import pandas as pd
from openpyxl import Workbook, load_workbook
from pubman_manager import PUBLICATIONS_DIR, PubmanCreator
from pubman_manager.excel_generator import create_sheet

def test_create_publications(monkeypatch, mock_calls_dir):
    def mock_create_items(self, request_list, create_items = True, submit_items=False, overwrite=False):
        print("request_list", request_list)

    monkeypatch.setattr(
        PubmanCreator,
        "create_items",
        mock_create_items,
        raising=True,
    )
    pc = PubmanCreator()

    pc.create_publications(mock_calls_dir / 'test_pubman_creator_cone_journal_misspelled.xlsx')


def test_create_talks_accepts_yes_no_invited(monkeypatch, tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MainSheet"
    headers = [
        "Event Name",
        "Conference start date\n(dd.mm.YYYY)",
        "Conference end date\n(dd.mm.YYYY)",
        "Talk date\n(dd.mm.YYYY)",
        "Conference Location\n(City, Country)",
        "Invited (yes/no)",
        "Type (Talk/Poster)",
        "Talk Title",
        "Author 1",
        "Affiliation 1",
    ]
    sheet.append(headers)
    sheet.append([
        "deRSE23 - Conference for Research Software Engineering in Germany",
        "20.03.2023",
        "22.03.2023",
        "21.03.2023",
        "Paderborn, Germany",
        "no",
        "Talk",
        "DAMASK: Challenges in collaborative development and outlook",
        "Ada Lovelace",
        "Analytical Engine Institute",
    ])
    sheet.append([
        "Conference X",
        "20.03.2023",
        "22.03.2023",
        "21.03.2023",
        "Paderborn, Germany",
        "yes",
        "Talk",
        "My Talk",
        "Ada Lovelace",
        "Analytical Engine Institute",
    ])
    file_path = tmp_path / "talks_yes_no.xlsx"
    workbook.save(file_path)

    creator = PubmanCreator.__new__(PubmanCreator)
    creator.ctx_id = "ctx"
    creator.user_id = "user_1"
    creator.identifier_paths = {}
    creator.authors_info = {}
    creator.journals = {}

    captured = {}

    def fake_create_items(request_list, create_items=True, submit_items=False, overwrite=False):
        captured["request_list"] = request_list
        return {"created": len(request_list), "total": len(request_list)}

    monkeypatch.setattr(creator, "create_items", fake_create_items)

    summary = creator.create_talks(file_path)

    assert summary == {"created": 1, "total": 1}
    request = captured["request_list"][0][1]
    assert request["metadata"]["event"]["invitationStatus"] == "INVITED"
    assert request["files"] == []


def test_create_talks_adds_external_link_file(monkeypatch, tmp_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "MainSheet"
    headers = [
        "Event Name",
        "Conference start date\n(dd.mm.YYYY)",
        "Conference end date\n(dd.mm.YYYY)",
        "Talk date\n(dd.mm.YYYY)",
        "Conference Location\n(City, Country)",
        "Invited (yes/no)",
        "Type (Talk/Poster)",
        "Talk Title",
        "External Link (optional)",
        "Author 1",
        "Affiliation 1",
    ]
    sheet.append(headers)
    sheet.append([
        "deRSE23 - Conference for Research Software Engineering in Germany",
        "20.03.2023",
        "22.03.2023",
        "21.03.2023",
        "Paderborn, Germany",
        "no",
        "Talk",
        "DAMASK: Challenges in collaborative development and outlook",
        "",
        "Ada Lovelace",
        "Analytical Engine Institute",
    ])
    external_url = "https://journals.aps.org/prl/abstract/10.1103/PhysRevLett.117.224101"
    sheet.append([
        "Conference X",
        "20.03.2023",
        "22.03.2023",
        "21.03.2023",
        "Paderborn, Germany",
        "no",
        "Talk",
        "My Talk",
        external_url,
        "Ada Lovelace",
        "Analytical Engine Institute",
    ])
    file_path = tmp_path / "talks_external_link.xlsx"
    workbook.save(file_path)

    creator = PubmanCreator.__new__(PubmanCreator)
    creator.ctx_id = "ctx"
    creator.user_id = "user_1"
    creator.identifier_paths = {}
    creator.authors_info = {}
    creator.journals = {}

    captured = {}

    def fake_create_items(request_list, create_items=True, submit_items=False, overwrite=False):
        captured["request_list"] = request_list
        return {"created": len(request_list), "total": len(request_list)}

    monkeypatch.setattr(creator, "create_items", fake_create_items)

    summary = creator.create_talks(file_path)

    assert summary == {"created": 1, "total": 1}
    request = captured["request_list"][0][1]
    assert request["files"] == [{
        "objectId": "",
        "lastModificationDate": "",
        "creationDate": "",
        "creator": {"objectId": ""},
        "visibility": "PUBLIC",
        "content": external_url,
        "storage": "EXTERNAL_URL",
        "size": 0,
        "metadata": {
            "title": external_url,
            "contentCategory": "publisher-version",
            "size": 0,
        },
    }]


def test_create_sheet_adds_yes_no_validation(tmp_path):
    file_path = tmp_path / "talks_template.xlsx"
    create_sheet(
        file_path,
        {("Ada", "Lovelace"): {"Analytical Engine Institute": 1}},
        OrderedDict([
            ("Event Name", [35, ""]),
            ("Invited (yes/no)", [15, "Select yes or no"]),
            ("Type (Talk/Poster)", [15, ""]),
            ("Talk Title", [50, ""]),
            ("External Link (optional)", [50, ""]),
        ]),
        n_authors=1,
        header_name="Event Name",
        n_entries=1,
        example_row=["Example", "no", "Talk", "Sample title", "", "Ada Lovelace", "Analytical Engine Institute"],
        freeze_first_n_cols=0,
    )

    workbook = load_workbook(file_path)
    sheet = workbook["MainSheet"]
    validations = list(sheet.data_validations.dataValidation)

    assert any(
        validation.formula1 == '"yes,no"'
        for validation in validations
    )

    header_row = None
    for row in sheet.iter_rows():
        if row[0].value == "Event Name":
            header_row = row
            break

    assert header_row is not None
    assert header_row[0].value == "Event Name"
    assert sheet.freeze_panes == "A14"

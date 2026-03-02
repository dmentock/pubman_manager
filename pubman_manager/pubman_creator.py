import json
import logging
import math
import re
import requests

from collections import OrderedDict
from datetime import datetime, date, timedelta
from dateutil import parser
from typing import Any
from pathlib import Path
from openpyxl import load_workbook

from pubman_manager import PubmanBase, FILES_DIR
from pubman_manager import get_user_cache_dir
from pubman_manager.util import is_mpi_affiliation, load_yaml


logger = logging.getLogger(__name__)


class PubmanCreator(PubmanBase):
    """
    PubmanCreator with unified Excel parsing:
    - pandas removed
    - openpyxl-only
    - extract_prefilled_rows() is the single parser
    """

    def __init__(self, base_url="https://pure.mpg.de/rest", auth_token=None, user_id=None):
        super().__init__(auth_token=auth_token, user_id=user_id, base_url=base_url)

        cache_dir = get_user_cache_dir(self.user_id)
        self.identifier_paths = load_yaml(cache_dir / 'identifier_paths.yaml')
        self.authors_info = load_yaml(cache_dir / 'authors_info.yaml')
        self.journals = load_yaml(cache_dir / 'journals.yaml')


    @staticmethod
    def extract_prefilled_rows(file_obj, header_name: str = "Title", limit: int | None = None):
        """
        Unified Excel parser:
        - Finds header row containing a cell == header_name (case-insensitive)
        - Deduplicates header column names
        - Reads all rows beneath header
        - Cleans cell content (strip, replace nan/None, unicode fixes)
        - Returns list of dictionaries: [{col: value, ...}, ...]

        """

        # Load Excel workbook
        stream = file_obj
        reset_stream = hasattr(stream, "seek")
        if reset_stream:
            current_pos = stream.tell()
            stream.seek(0)

        workbook = load_workbook(stream, read_only=True, data_only=True)
        try:
            # Use first sheet unless "MainSheet" exists
            sheet = workbook["MainSheet"] if "MainSheet" in workbook.sheetnames else workbook[workbook.sheetnames[0]]

            header_row_idx = None
            raw_header = None

            # ------------------------------
            # 1. Locate header row
            # ------------------------------
            for idx, row in enumerate(sheet.iter_rows(values_only=True)):
                if not row:
                    continue
                for cell in row:
                    if isinstance(cell, str) and cell.strip().casefold() == header_name.casefold():
                        header_row_idx = idx
                        raw_header = list(row)
                        break
                if header_row_idx is not None:
                    break

            if header_row_idx is None:
                raise ValueError(f"Header row with '{header_name}' not found.")

            # ------------------------------
            # 2. Clean + dedupe header names
            # ------------------------------
            header = []
            seen = {}

            def clean_colname(x):
                if x is None:
                    return ""
                s = str(x).strip()
                return "" if s.lower() in {"nan", "none"} else s

            for i, h in enumerate(raw_header):
                name = clean_colname(h)
                if not name:
                    name = f"Unnamed_{i}"

                if name in seen:
                    seen[name] += 1
                    name = f"{name}_{seen[name]}"
                else:
                    seen[name] = 0

                header.append(name)

            # ------------------------------
            # 3. Read data rows beneath the header
            # ------------------------------
            data_rows = []

            for row in sheet.iter_rows(min_row=header_row_idx + 2, values_only=True):
                if row is None:
                    continue

                # Clean row cells
                cleaned = []
                row_has_data = False

                for cell in row[: len(header)]:
                    if cell is None:
                        cleaned.append("")
                        continue

                    s = str(cell).strip()
                    if s.lower() in {"nan", "none"}:
                        cleaned.append("")
                        continue

                    cleaned.append(s)
                    if s != "":
                        row_has_data = True

                if not row_has_data:
                    continue

                row_dict = {header[i]: cleaned[i] for i in range(len(cleaned))}
                data_rows.append(row_dict)

                if limit is not None and len(data_rows) >= limit:
                    break

            return data_rows

        finally:
            workbook.close()
            if reset_stream:
                stream.seek(current_pos)

    # -------------------------------------------------------------------------
    #  AUTHOR PARSING / CLEANUP UTILITIES
    # -------------------------------------------------------------------------

    def get_row_authors_info(self, row):
        """
        Works exactly like before — row is now a dict generated from the unified parser.
        """
        xnnnn_re = re.compile(r"_x([0-9A-Fa-f]{4})_")

        row_authors = OrderedDict()
        authors_info_merged_names = {
            f'{first_name} {last_name}': affiliations
            for (first_name, last_name), affiliations in self.authors_info.items()
        }

        for i in range(1, 50):
            author_name_key = f'Author {i}'
            affiliation_key = f'Affiliation {i}'

            if author_name_key not in row or affiliation_key not in row:
                continue

            author = row.get(author_name_key)
            affiliation = row.get(affiliation_key)

            if not author or not affiliation:
                continue

            if author not in row_authors:
                if identifier := authors_info_merged_names.get(author, {}).get('identifier'):
                    row_authors[author] = {'identifier': identifier}
                else:
                    row_authors[author] = {}

            cleaned_aff = xnnnn_re.sub(lambda m: chr(int(m.group(1), 16)), affiliation).replace('\r', '')
            row_authors[author].setdefault('affiliations', []).append(cleaned_aff)

        return row_authors

    # -------------------------------------------------------------------------
    #  DATE UTILITIES
    # -------------------------------------------------------------------------

    def safe_date_parse(self, val):
        """
        Same implementation as before.
        Accepts: datetime, date, serial numbers, various date formats.
        """
        if val is None or val == "":
            return None

        if isinstance(val, datetime):
            return val.replace(hour=0, minute=0, second=0, microsecond=0)
        if isinstance(val, date):
            return datetime(val.year, val.month, val.day)

        if isinstance(val, (int, float)) and not (isinstance(val, float) and math.isnan(val)):
            if val > 59:
                base = datetime(1899, 12, 30)
                return base + timedelta(days=int(val))

        s = str(val).strip().replace('\u200e', '').replace('\u200f', '')

        if re.fullmatch(r'\d{4}-\d{2}-\d{2}', s):
            return datetime.strptime(s, '%Y-%m-%d')

        if re.fullmatch(r'\d{1,2}\.\d{1,2}\.\d{4}', s):
            return datetime.strptime(s, '%d.%m.%Y')

        if re.fullmatch(r'\d{1,2}/\d{1,2}/\d{4}', s):
            m, d, y = map(int, s.split('/'))
            return datetime(y, m, d)

        m = re.search(r'(?<!\d)(\d{1,2})[.\-/](\d{4})(?!\d)', s)
        if m:
            mth, yr = int(m.group(1)), int(m.group(2))
            if 1 <= mth <= 12:
                return datetime(yr, mth, 1)

        m = re.search(r'(?<!\d)(\d{4})-(\d{1,2})(?!\d)', s)
        if m:
            yr, mth = int(m.group(1)), int(m.group(2))
            if 1 <= mth <= 12:
                return datetime(yr, mth, 1)

        dt = parser.parse(s, dayfirst=False, yearfirst=False, default=datetime(1900, 1, 1))
        return dt.replace(hour=0, minute=0, second=0, microsecond=0)

    def format_date(self, parsed_date, original):
        if not parsed_date:
            return None
        if "-" in original:
            parts = original.split("-")
        else:
            parts = original.split(".")

        if len(parts) == 1:
            return parsed_date.strftime("%Y")
        if len(parts) == 2:
            return parsed_date.strftime("%Y-%m")
        return parsed_date.strftime("%Y-%m-%d")

    # -------------------------------------------------------------------------
    #  AFFILIATION CLEANUP
    # -------------------------------------------------------------------------

    def clean_affiliation(self, affiliation):
        return re.sub(r'_x[0-9A-Fa-f]+_', '', affiliation)

    def clean_scalar(self, scalar):
        if not scalar:
            return ''
        try:
            return str(int(scalar))
        except:
            return str(scalar)

    # -------------------------------------------------------------------------
    #  PDF UPLOAD
    # -------------------------------------------------------------------------

    def upload_pdf(self, pdf_path):
        pdf_path = Path(pdf_path)
        staging_url = f"{self.base_url}/staging/{pdf_path.name}"

        headers = {"Authorization": self.auth_token, "Content-Type": "application/pdf"}

        with open(pdf_path, 'rb') as file_data:
            payload = file_data.read()
            response = requests.post(staging_url, headers=headers, data=payload)

        if response.status_code in [200, 201]:
            return response.json()
        else:
            raise Exception(f"Failed to upload file: {response.status_code} {response.text}")

    def get_first_and_last_name_from_concat(self, name):
        parts = name.split()
        given = []
        family = []
        for i, part in enumerate(parts):
            if part.endswith('.') or i == 0:
                given.append(part)
            else:
                family.append(part)
        return " ".join(given), " ".join(family)

    def get_journal_by_issn(self, issn: str) -> dict | None:
        query = {
            "query": {
                "nested": {
                    "path": "metadata.sources.identifiers",
                    "query": {
                        "bool": {
                            "must": [{"match_phrase": {"metadata.sources.identifiers.type": "ISSN"}}],
                            "should": [{"match_phrase": {"metadata.sources.identifiers.id": issn}}],
                            "minimum_should_match": 1,
                        }
                    }
                }
            },
            "size": 25,
        }

        headers = {"Authorization": self.auth_token, "Content-Type": "application/json"}

        resp = requests.post(f"{self.base_url}/items/search", headers=headers, data=json.dumps(query))
        if resp.status_code != 200:
            raise Exception(f"Journal lookup failed: {resp.status_code} {resp.text}")

        records = resp.json().get("records", []) or []

        def pick_source(sources):
            matches = []
            for s in sources or []:
                ids = {i.get("type", "").upper(): i.get("id") for i in s.get("identifiers", [])}
                if "ISSN" not in ids:
                    continue
                matches.append(("CONE" in ids, s, ids))
            if not matches:
                return None
            matches.sort(key=lambda x: (not x[0]))
            _, source, ids = matches[0]
            return {
                "title": source.get("title"),
                "alternativeTitles": source.get("alternativeTitles", []),
                "genre": source.get("genre"),
                "publishingInfo": source.get("publishingInfo"),
                "cone": ids.get("CONE"),
            }

        for rec in records:
            srcs = rec.get("data", {}).get("metadata", {}).get("sources", [])
            best = pick_source(srcs)
            if best:
                return best
        return None

    def create_talks(self, file_path, create_items=True, submit_items=False, overwrite=False):
        rows = self.extract_prefilled_rows(file_path, header_name="Event Name")
        example_index = None
        for idx, row in enumerate(rows):
            for value in row.values():
                if isinstance(value, str) and value.strip().casefold() == "example":
                    example_index = idx
                    break
            if example_index is not None:
                break
        if example_index is not None:
            rows = rows[example_index + 1 :]
        request_list = []
        missing_pdfs = []

        for row in rows:
            logger.info(f'Generating requests for "{row.get("Talk Title")}"')

            if not row.get("Talk Title"):
                continue

            row_authors_info = self.get_row_authors_info(row)

            metadata_creators = []
            for author, info in row_authors_info.items():
                given, family = self.get_first_and_last_name_from_concat(author)

                affiliation_list = []
                for aff in info['affiliations']:
                    entry = {"name": self.clean_affiliation(aff), "identifierPath": [""]}

                    if aff in self.identifier_paths:
                        entry["identifier"] = self.identifier_paths[aff][0]
                    elif not is_mpi_affiliation(aff):
                        entry["identifier"] = 'ou_persistent22'

                    affiliation_list.append(entry)

                metadata_creators.append({
                    "person": {
                        "givenName": given,
                        "familyName": family,
                        "organizations": affiliation_list,
                        "identifier": info.get('identifier'),
                    },
                    "role": "AUTHOR",
                    "type": "PERSON",
                })

            talk_type = row.get('Type (Talk/Poster)', '').lower().strip()
            if talk_type == 'talk':
                genre = "TALK_AT_EVENT"
            elif talk_type == 'poster':
                genre = "POSTER"
            else:
                raise RuntimeError(f'Invalid Type (Talk/Poster): "{row.get("Type (Talk/Poster)")}"')

            request = {
                "context": {"objectId": self.ctx_id, "name": "", "lastModificationDate": "",
                            "creationDate": "", "creator": {"objectId": ""}},
                "creator": {"objectId": self.user_id},
                "modifier": {"objectId": self.user_id},
                "localTags": [],
                "metadata": {
                    "title": row.get('Talk Title'),
                    "creators": metadata_creators,
                    "dateCreated": self.safe_date_parse(row.get('Talk date\n(dd.mm.YYYY)')).strftime("%Y-%m-%d")
                        if row.get('Talk date\n(dd.mm.YYYY)') else None,
                    "genre": genre,
                    "event": {
                        "endDate": self.safe_date_parse(row.get('Conference end date\n(dd.mm.YYYY)')).strftime("%Y-%m-%d")
                            if row.get('Conference end date\n(dd.mm.YYYY)') and
                               row.get('Conference end date\n(dd.mm.YYYY)') != row.get('Conference start date\n(dd.mm.YYYY)')
                            else None,
                        "place": row.get('Conference Location\n(City, Country)'),
                        "startDate": self.safe_date_parse(row.get('Conference start date\n(dd.mm.YYYY)')).strftime("%Y-%m-%d"),
                        "title": row.get('Event Name'),
                        "invitationStatus": "INVITED" if row.get('Invited (y/n)', "").strip().lower() == 'y' else None
                    },
                    "languages": ["eng"],
                },
                "files": [],
            }

            request_list.append((
                {"metadata.title": row.get('Talk Title'), "metadata.event.title": row.get('Event Name')},
                request
            ))

        if not request_list:
            raise RuntimeError("No talk rows found after the example row.")

        summary = self.create_items(
            request_list,
            create_items=create_items,
            submit_items=submit_items,
            overwrite=overwrite,
        )
        return summary

    # -------------------------------------------------------------------------
    #  CREATE PUBLICATIONS (now using unified parser)
    # -------------------------------------------------------------------------

    def create_publications(self, file_path, submit_items=False, overwrite=False):
        rows = self.extract_prefilled_rows(file_path, header_name="Title")
        request_list = []
        missing_pdfs = []

        if not rows:
            raise RuntimeError("No publication rows found in the uploaded file.")

        for row in rows:
            title = row.get("Title")
            if not title:
                raise RuntimeError(f"Missing title in row: {row}")

            row_authors_info = self.get_row_authors_info(row)
            metadata_creators = []

            for author, info in row_authors_info.items():
                given, family = self.get_first_and_last_name_from_concat(author)

                affiliation_list = []
                for aff in info['affiliations']:
                    if aff in self.identifier_paths:
                        affiliation_list.append({
                            "name": aff,
                            "identifier": self.identifier_paths[aff][0],
                            "identifierPath": [""],
                        })
                    else:
                        identifier = 'ou_persistent22'
                        affiliation_list.append({
                            "name": aff,
                            "identifier": identifier,
                            "identifierPath": [""],
                        })

                metadata_creators.append({
                    "person": {
                        "givenName": given.strip(),
                        "familyName": family.strip(),
                        "organizations": affiliation_list,
                        "identifier": info.get('identifier'),
                    },
                    "role": "AUTHOR",
                    "type": "PERSON",
                })

            # Dates
            doi = str(row.get("DOI"))

            date_issued_raw = row.get("Date issued", "")
            date_issued_parsed = self.safe_date_parse(date_issued_raw) if date_issued_raw else None
            date_issued = self.format_date(date_issued_parsed, date_issued_raw) if date_issued_parsed else None

            date_online_raw = row.get("Date published online", "")
            date_online_parsed = self.safe_date_parse(date_online_raw) if date_online_raw else None
            date_online = self.format_date(date_online_parsed, date_online_raw) if date_online_parsed else None

            # Journal
            issn = row.get("ISSN")
            journal_title = row.get("Journal Title")

            if issn in self.journals:
                journal_info = self.journals.get(issn)
            else:
                logger.info(f'No local journal entry for "{journal_title}" ({issn}), looking globally...')
                journal_info = self.get_journal_by_issn(issn)
                if not journal_info:
                    logger.warning(f'No CoNe entry found for journal {journal_title} ({issn})')
                    journal_info = {}

            sources = [{
                'alternativeTitles': journal_info.get('alternativeTitles', []),
                'genre': journal_info.get('genre', 'JOURNAL'),
                'title': journal_title,
                'publishingInfo': journal_info.get('publishingInfo', {'publisher': row.get('Publisher')}),
                'volume': self.clean_scalar(row.get('Volume')),
                'issue': self.clean_scalar(row.get('Issue')),
                'identifiers': [
                    {'type': 'ISSN', 'id': issn},
                    {'type': 'CONE', 'id': journal_info.get('cone')},
                ],
            }]

            # Pages
            page = self.clean_scalar(row.get('Page'))
            if page:
                if "-" in page:
                    p1, p2 = page.split("-", 1)[0].strip(), page.split("-", 1)[1].strip()
                else:
                    p1, p2 = page.strip(), ""

                sources[0]['startPage'] = p1
                if p2:
                    sources[0]['endPage'] = p2

                m1 = re.match(r"^\s*([A-Za-z]*)(\d+)\s*$", p1)
                m2 = re.match(r"^\s*([A-Za-z]*)(\d+)\s*$", p2) if p2 else None
                if m1 and m2:
                    prefix1, n1 = m1.group(1).lower(), int(m1.group(2))
                    prefix2, n2 = m2.group(1).lower(), int(m2.group(2))
                    if prefix1 == prefix2 and n2 >= n1:
                        sources[0]['totalNumberOfPages'] = n2 - n1 + 1

            # Article number
            article_number = self.clean_scalar(row.get("Article Number"))
            if article_number:
                sources[0]['sequenceNumber'] = article_number

            # PDF
            files = []
            pdf_path = Path(FILES_DIR / f'{doi.replace("/", "")}.pdf')

            license_url = row.get('License url')
            if license_url:
                if not pdf_path.exists():
                    logger.error(f'PDF for DOI {doi} not found: {pdf_path}')
                    missing_pdfs.append(pdf_path.name)
                else:
                    file_id = self.upload_pdf(pdf_path)
                    file_entry = {
                        "objectId": '',
                        "name": pdf_path.name,
                        "lastModificationDate": "",
                        "creationDate": "",
                        "creator": {"objectId": ""},
                        "pid": "",
                        'content': file_id,
                        "visibility": "PUBLIC",
                        "contentCategory": "publisher-version",
                        "checksum": "",
                        "checksumAlgorithm": "MD5",
                        "mimeType": "",
                        "size": 0,
                        'storage': 'INTERNAL_MANAGED',
                        "metadata": {
                            "title": pdf_path.name,
                            "description": "File downloaded from scopus",
                            "formats": [{"value": "", "type": ""}],
                            "size": 0,
                            "license": license_url,
                        },
                    }
                    license_year = row.get("License year")
                    if license_year:
                        try:
                            file_entry["metadata"]["copyrightDate"] = str(int(license_year))
                        except:
                            pass
                    if 'arxiv' in license_url.lower():
                        file_entry["metadata"]["contentCategory"] = "pre-print"
                    files.append(file_entry)

            # Build final request
            request = {
                "context": {"objectId": self.ctx_id, "name": "", "lastModificationDate": "",
                            "creationDate": "", "creator": {"objectId": self.user_id}},
                "creator": {"objectId": self.user_id},
                "modifier": {"objectId": self.user_id},
                "localTags": [],
                "metadata": {
                    "title": title,
                    "creators": metadata_creators,
                    "datePublishedInPrint": date_issued,
                    "datePublishedOnline": date_online,
                    "genre": "ARTICLE",
                    "identifiers": [{"id": doi, "type": "DOI"}],
                    "languages": ["eng"],
                    "sources": sources,
                    "reviewMethod": "PEER",
                },
                "files": files,
            }

            criteria = {"metadata.identifiers": {"id": doi, "type": "DOI"}}
            request_list.append((criteria, request))

        summary = self.create_items(request_list, submit_items=submit_items, overwrite=overwrite)
        if missing_pdfs:
            raise RuntimeError(
                "Could not find PDF(s) for open access publications: "
                + ", ".join(sorted(set(missing_pdfs)))
            )
        return summary

    def create_items(self, request_list, create_items=True, submit_items=False, overwrite=False):
        item_ids = []
        created_count = 0
        skipped_existing = 0
        blocked_existing = 0

        for criteria, request_json in request_list:
            created_item = None
            title = request_json['metadata']['title']
            existing = self.search_publication_by_criteria(criteria)

            if existing:
                if overwrite:
                    logger.info(f"Overwriting existing publication: '{title}'")
                    item_already_released = False
                    for pub in existing:
                        deleted = self.delete_item(pub['data']['objectId'], pub['data']['lastModificationDate'])
                        if not deleted:
                            item_already_released = True
                            logger.info(f"Could not delete publication '{title}', skipping")
                    if item_already_released:
                        blocked_existing += 1
                        continue
                else:
                    logger.info(f"Skipping existing publication: '{criteria}'")
                    pub = existing[0]['data']
                    item_ids.append((pub['objectId'], pub['lastModificationDate'], pub['versionState']))
                    skipped_existing += 1
                    continue

            if create_items:
                created_item = self.create_item(request_json)
                if created_item:
                    item_ids.append((created_item['objectId'], created_item['lastModificationDate'],
                                     created_item['versionState']))
                    created_count += 1

        if submit_items:
            for obj_id, mod, state in item_ids:
                if state not in ['PENDING', 'IN_REVISION']:
                    logger.info(f"Item already in state '{state}', skipping submit")
                    continue
                submitted = self.submit_item(obj_id, mod)
                logger.info(f"Submitted item: {submitted}")
        return {
            "created": created_count,
            "skipped_existing": skipped_existing,
            "blocked_existing": blocked_existing,
            "total": len(request_list),
        }
